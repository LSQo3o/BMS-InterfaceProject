"""
从excel中读取数据文件
得到测试数据格式为[[],[]],做为从参数传入pytest中
"""
import openpyxl

from Conf.VarConfig import BASE_DIR


def get_row_values(sheet_name, row):
    # 得到某一列的值，以列表形式返回
    file_path = BASE_DIR + "/Data" + "/InterFaceData.xlsx"
    work_book = openpyxl.load_workbook(file_path)
    sheet = work_book[sheet_name]
    data = []
    columns = sheet.max_column  # 获取最大列数
    for i in range(1, columns + 1):
        data.append(sheet.cell(row=row, column=i).value)
    return data
    # print( data)


def get_cell_values(sheet_name, row, column):
    # 得到某一单元格的值，以列表的形式返回
    file_path = BASE_DIR + "/Data" + "/InterFaceData.xlsx"
    work_book = openpyxl.load_workbook(file_path)
    sheet = work_book[sheet_name]
    data = []
    value = sheet.cell(row=row, column=column).value
    data.append(value)
    return data


def get_test_data(sheet_name):
    # 得到测试数据的值，以[[],[]]的形式返回
    file_path = BASE_DIR + "/Data" + "/InterFaceData.xlsx"
    work_book = openpyxl.load_workbook(file_path)
    sheet = work_book[sheet_name]
    data = []
    rows = sheet.max_row  # 获取最大行数
    columns = sheet.max_column  # 获取最大列数
    for i in range(2, rows + 1):
        sing_row_data = []
        for j in range(1, columns + 1):
            value = sheet.cell(row=i, column=j).value
            sing_row_data.append(value)
        data.append(sing_row_data)
    return data


if __name__ == '__main__':
    get_row_values("login", 1)
    get_test_data("login")
